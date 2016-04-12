#!/usr/bin/python
# -*- coding:utf-8 -*-
import os.path

try:
    from openpyxl.worksheet import worksheet
    from openpyxl import load_workbook, workbook
except ImportError as e:
    print("you need install openpyxl module!")
    print("Please use command:")
    print("  pip install openpyxl")
    raise


def ReadLine(ws):
    """
    逐行读取工作簿中的内容
    """
    if not isinstance(ws, worksheet.Worksheet):
        raise TypeError("parament type is "
                        "openpyxl.worksheet.worksheet.Worksheet")
    
    for row in ws.rows:
        yield [x.value for x in row]


def copysheet(wb_x_filename, wb_y_filename, sheetname, forceflg=False):
    """
    将wb_x_filename中的工作簿sheetname拷贝到wb_y_filename中。如果wb_y_filename
    中存在sheetname工作簿,会根据标志位forceflg判断是否强制覆盖
    :param wb_x_filename: str 文件A
    :param wb_y_filename: str 文件B
    :param sheetname: str 待拷贝工作簿名,需在文件A中存在
    :param forceflg: bool 是否强制覆盖标志。如果文件B中存在同名工作簿时,在
                     forceflg为True时会覆盖
    :return:
        True -- 拷贝成功
        False -- 未覆盖文件B中sheetname工作簿中内容
    """
    if not os.path.exists(wb_x_filename):
        raise FileNotFoundError("Workbook:%s not found" % wb_x_filename)

    if not os.path.exists(wb_y_filename):
        raise FileNotFoundError("Workbook:%s not found" % wb_y_filename)

    if not isinstance(sheetname, str):
        raise TypeError("sheetname must be str !")

    wb_x = load_workbook(wb_x_filename)
    wb_y = load_workbook(wb_y_filename)

    if sheetname not in wb_x:
        raise KeyError("Worksheet %s not in wb_x")

    if forceflg not in (True, False):
        raise ValueError("forceflg's value range need in (True,False)")

    ws_x = wb_x.get_sheet_by_name(sheetname)

    if sheetname in wb_y and forceflg:
        ws_y = wb_y.get_sheet_by_name(sheetname)
    elif sheetname not in wb_y:
        ws_y = wb_y.create_sheet(title=sheetname)
    else:
        return False

    for xrow in ws_x.rows:
        for xcell in xrow:
            ws_y[xcell.coordinate] = xcell.value

    wb_y.save(wb_y_filename)
    return True

if __name__ == '__main__':

    wb = load_workbook('test.xlsx')
    ws = wb.active
    for line in ReadLine(ws):
        print(line)

    mypath = r'D:\5.sparetime_program\Python\Test\\'
    test1 = mypath + 'test1.xlsx'
    test2 = mypath + 'test2.xlsx'
    if copysheet(test1, test2, 'mytest', True):
        print("Copy Suc!")
